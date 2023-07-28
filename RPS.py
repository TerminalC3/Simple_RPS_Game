import random
import openpyxl as xl

score = 0
wins = 0
loses = 0
total_rounds = 0
draws = 0
rating = 0
username = 'Default'
record_code = 0


def high():
    global score
    global wins
    global loses
    global draws
    global total_rounds
    global rating
    global username

    if total_rounds != 0:
        rating = (int(total_rounds) / int(score)) * 100

    else:
        rating = 0

    wb = xl.load_workbook('high scores.xlsx')
    sheet = wb['Sheet1']

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 1)

        if cell.value == None:
            cell.value = username
            break

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 2)

        if cell.value == None:
            cell.value = score
            break

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 3)

        if cell.value == None:
            cell.value = wins
            break

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 4)

        if cell.value == None:
            cell.value = loses
            break

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 5)

        if cell.value == None:
            cell.value = draws
            break

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 6)

        if cell.value == None:
            cell.value = total_rounds
            break

    for row in range(1, 101):
        cell = sheet.cell(row + 1, 7)

        if cell.value == None:
            cell.value = rating
            break

    wb.save('high scores.xlsx')
    sortit()


ro = 1
max_rowy = 0


def sortit():
    wb = xl.load_workbook('high scores.xlsx')
    sheet = wb['Sheet1']

    def sorthigh():
        # for number in range(1, 101):
        #     for row in range(1, 8):
        #         for col in range(100, 201):
        #             sheet.cell(row, col).value = sheet.cell(number, row).value
        for u in range(1, 101):
            max_value = 0
            global max_rowy
            # ended = False
            for row in range(2, 101):
                if sheet.cell(row, 7).value == None:
                    pass

                elif sheet.cell(row, 7).value == '':
                    pass

                # elif sheet.cell(row, 7).value == 'end':
                #     ended = True

                elif float(sheet.cell(row, 7).value) > float(max_value):
                    max_value = sheet.cell(row, 7).value
                    max_rowy = int(row)

                else:
                    pass

            i = 1
            global ro
            ro = ro + 1

            for f in range(8, 15):
                sheet.cell(ro, f).value = sheet.cell(max_rowy, i).value
                i += 1

            for number in range(1, 8):
                sheet.cell(max_rowy, number).value = ''

            wb.save('high scores.xlsx')

            # if sheet.cell(9, 14).value != None:
            #     exit()
            #
            # elif ended == True:
            #     sorthigh()

    sorthigh()

    for danumber in range(2, 101):
        numerals = 8
        for numbers in range(1, 8):
            sheet.cell(danumber, numbers).value = sheet.cell(danumber, numerals).value
            numerals += 1

    for danumber in range(2, 101):
        for numbers in range(8, 15):
            sheet.cell(danumber, numbers).value = ''

    wb.save('high scores.xlsx')


def choose():
    global score
    global wins
    global loses
    global total_rounds
    global draws
    global rating

    if total_rounds != 0:
        rating = (int(total_rounds) / int(score)) * 100

    else:
        rating = 0

    print('''
    s   -   Start/Resume
    r   -   Reset
    q   -   Quit
    sc  -   Show scoreboard
    rec -   Record your score

    IN THE MIDDLE OF A GAME
    o - options
    ''')

    option = input(">>")

    if option.lower() == 'q':
        print("Shutting down...")
        exit()

    elif option.upper() == 'S':
        main()

    elif option.upper() == 'SC':
        print('Score = ', score)
        print('Wins = ', wins)
        print('Loses = ', loses)
        print('Draws =', draws)
        print('Total rounds =', total_rounds)
        print('Rating =', rating)
        choose()

    elif option.upper() == 'REC':
        global username
        global record_code

        if rating > 0:
            if record_code == 0:
                username = input("Recording your score \n Enter your username: ")
                high()
                print("Your scores have been recorded")
                record_code = 1
                main()

            else:
                high()
                print("Your scores have been recorded")
                main()

        else:
            print('Ratings lower than 0 can not be recorded')
            main()

    elif option.lower() == 'r':
        confirm = input('Are you sure you want to reset your score? You will lose all progress (y/n): ')
        if confirm.upper() == 'Y' or confirm.upper() == "YES":
            score = 0
            wins = 0
            loses = 0
            total_rounds = 0
            draws = 0
            record_code = 0

            print('Score has been reset')
            main()

        elif confirm.upper() == 'N' or confirm.upper() == 'NO':
            print('Continuing')
            main()

        else:
            print('Invalid option!')

    else:
        print("Invalid option! Please refer above for available options")
        choose()


def main():
    global total_rounds
    global score
    global wins
    global loses
    global draws

    choice = random.randint(1, 3)
    if choice == 1:
        opp_choice = 'Rock'

    elif choice == 2:
        opp_choice = 'Paper'

    elif choice == 3:
        opp_choice = 'Scissors'

    user_choice = input("\nYour choice: ")
    if user_choice.upper() == 'R' or user_choice.upper() == 'ROCK':
        user_choice = 'Rock'

    elif user_choice.upper() == 'P' or user_choice.upper() == 'PAPER':
        user_choice = 'Paper'

    elif user_choice.upper() == 'S' or user_choice.upper() == 'SCISSORS':
        user_choice = "Scissors"

    elif user_choice.upper() == 'O' or user_choice.upper() == 'OPTIONS':
        choose()

    else:
        print("Invalid option!")
        main()

    if user_choice == opp_choice:
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("Its a draw!")
        score += 25
        total_rounds += 1
        draws += 1
        main()

    elif user_choice == 'Rock' and opp_choice == 'Scissors':
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("You won!")
        score += 100
        total_rounds += 1
        wins += 1
        main()

    elif user_choice == 'Paper' and opp_choice == 'Rock':
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("You won!")
        score += 100
        total_rounds += 1
        wins += 1
        main()

    elif user_choice == 'Scissors' and opp_choice == 'Paper':
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("You won!")
        score += 100
        total_rounds += 1
        wins += 1
        main()

    elif user_choice == 'Scissors' and opp_choice == 'Rock':
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("You lose!")
        score -= 75
        total_rounds += 1
        loses += 1
        main()

    elif user_choice == 'Paper' and opp_choice == 'Scissors':
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("You lose!")
        score -= 75
        total_rounds += 1
        loses += 1
        main()

    elif user_choice == 'Rock' and opp_choice == 'Paper':
        print("You chose", user_choice, ", I chose ", opp_choice)
        print("You lose!")
        score -= 75
        total_rounds += 1
        loses += 1
        main()


choose()

